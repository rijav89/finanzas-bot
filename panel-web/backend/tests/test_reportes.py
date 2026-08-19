"""Reportes agrupados y exportación.

Los archivos se verifican abriéndolos de vuelta, no mirando el tamaño: un Excel
corrupto también pesa varios kB.
"""
from datetime import date, timedelta
from io import BytesIO

import pytest

from tests.conftest import AUTH_UID_A, AUTH_UID_B, como

pytestmark = pytest.mark.asyncio

HOY = date.today()
MES = {"desde": HOY.replace(day=1).isoformat(), "hasta": HOY.isoformat()}


async def _sembrar(cliente, datos):
    """Dos categorías de gasto y un ingreso, todo en el mes en curso."""
    como(cliente, AUTH_UID_A)
    await cliente.post(
        "/api/v1/ingresos",
        json={
            "monto": "3000", "cuenta_id": datos["cuenta_a"],
            "categoria": "Sueldo", "descripcion": "sueldo de prueba",
        },
    )
    for monto, cat in (("200", "Comida"), ("50", "Comida"), ("120", "Vivienda")):
        await cliente.post(
            "/api/v1/gastos",
            json={"monto": monto, "categoria": cat, "cuenta_id": datos["cuenta_a"]},
        )


async def test_agrupar_por_categoria_suma_y_ordena(cliente, datos):
    await _sembrar(cliente, datos)

    d = (await cliente.get("/api/v1/reportes/resumen", params=MES)).json()["data"]

    por_clave = {f["clave"]: f for f in d["filas"]}
    assert por_clave["Comida"]["gastos"] == 250.0
    assert por_clave["Comida"]["n"] == 2
    assert por_clave["Vivienda"]["gastos"] == 120.0
    assert por_clave["Sueldo"]["ingresos"] == 3000.0
    assert por_clave["Sueldo"]["neto"] == 3000.0

    assert d["totales"] == {"gastos": 370.0, "ingresos": 3000.0, "neto": 2630.0, "n": 4}


async def test_filtro_por_tipo_deja_un_solo_lado(cliente, datos):
    await _sembrar(cliente, datos)

    d = (await cliente.get(
        "/api/v1/reportes/resumen", params={**MES, "tipo": "gasto"}
    )).json()["data"]

    assert d["totales"]["ingresos"] == 0.0
    assert d["totales"]["gastos"] == 370.0
    assert "Sueldo" not in {f["clave"] for f in d["filas"]}


async def test_agrupar_por_mes_y_por_cuenta(cliente, datos):
    await _sembrar(cliente, datos)

    d = (await cliente.get(
        "/api/v1/reportes/resumen", params={**MES, "group_by": "mes"}
    )).json()["data"]
    assert [f["clave"] for f in d["filas"]] == [f"{HOY:%Y-%m}"]
    assert d["filas"][0]["neto"] == 2630.0

    d = (await cliente.get(
        "/api/v1/reportes/resumen", params={**MES, "group_by": "cuenta"}
    )).json()["data"]
    assert [f["clave"] for f in d["filas"]] == ["Principal"]


async def test_prestamos_y_transferencias_fuera_del_reporte(cliente, datos):
    """Mismo criterio que el dashboard: mueven saldo, no son ingreso ni gasto."""
    await _sembrar(cliente, datos)
    await cliente.post(
        "/api/v1/transferencias",
        json={
            "origen_id": datos["cuenta_a"], "destino_id": datos["cuenta_a2"], "monto": "500",
        },
    )
    await cliente.post(
        "/api/v1/deudas",
        json={
            "tipo": "prestamo_recibido", "acreedor": "Juan", "monto_total": "800",
            "fecha_inicio": HOY.isoformat(), "cuenta_id": datos["cuenta_a"],
            "generar_cuotas": False,
        },
    )

    d = (await cliente.get("/api/v1/reportes/resumen", params=MES)).json()["data"]
    assert d["totales"] == {"gastos": 370.0, "ingresos": 3000.0, "neto": 2630.0, "n": 4}


async def test_rango_invertido_y_demasiado_largo(cliente, datos):
    como(cliente, AUTH_UID_A)
    r = await cliente.get(
        "/api/v1/reportes/resumen",
        params={"desde": HOY.isoformat(), "hasta": (HOY - timedelta(days=1)).isoformat()},
    )
    assert r.status_code == 400
    assert r.json()["error"] == "rango_invertido"

    r = await cliente.get(
        "/api/v1/reportes/resumen",
        params={"desde": (HOY - timedelta(days=4000)).isoformat(), "hasta": HOY.isoformat()},
    )
    assert r.status_code == 400
    assert r.json()["error"] == "rango_demasiado_largo"


async def test_no_ver_datos_de_otro_usuario(cliente, datos):
    await _sembrar(cliente, datos)

    como(cliente, AUTH_UID_B)
    d = (await cliente.get("/api/v1/reportes/resumen", params=MES)).json()["data"]
    assert d["filas"] == []
    assert d["totales"]["n"] == 0


# ── Exportación ──────────────────────────────────────────────────────────────

async def test_excel_se_abre_y_trae_las_dos_hojas(cliente, datos):
    from openpyxl import load_workbook

    await _sembrar(cliente, datos)
    r = await cliente.get("/api/v1/reportes/export.xlsx", params=MES)

    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment; filename=" in r.headers["content-disposition"]

    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames == ["Resumen", "Movimientos"]

    resumen = wb["Resumen"]
    fila_total = [c.value for c in resumen[resumen.max_row]]
    assert fila_total[0] == "TOTAL"
    assert fila_total[1] == 3000.0  # ingresos
    assert fila_total[2] == 370.0  # gastos

    # El detalle trae los 4 movimientos, con los gastos en negativo
    det = wb["Movimientos"]
    montos = [det.cell(row=f, column=6).value for f in range(2, det.max_row + 1)]
    assert sorted(montos) == [-200.0, -120.0, -50.0, 3000.0]


async def test_pdf_es_un_pdf_valido(cliente, datos):
    await _sembrar(cliente, datos)
    r = await cliente.get("/api/v1/reportes/export.pdf", params=MES)

    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content.startswith(b"%PDF-")
    assert r.content.rstrip().endswith(b"%%EOF")
    assert len(r.content) > 1000


async def test_export_de_un_periodo_vacio_no_falla(cliente, datos):
    """Sin movimientos el archivo igual tiene que salir, no reventar."""
    como(cliente, AUTH_UID_A)
    vacio = {"desde": "2020-01-01", "hasta": "2020-01-31"}

    assert (await cliente.get("/api/v1/reportes/export.xlsx", params=vacio)).status_code == 200
    r = await cliente.get("/api/v1/reportes/export.pdf", params=vacio)
    assert r.status_code == 200
    assert r.content.startswith(b"%PDF-")


# ── Un export a la vez por usuario ───────────────────────────────────────────

async def test_dos_exports_a_la_vez_del_mismo_usuario_rechaza_el_segundo(cliente, datos):
    """Pedir dos archivos juntos nunca es intencional: es doble clic o dos pestañas.
    Se rechaza en el acto en vez de hacer cola, que se siente como que se colgó."""
    import asyncio

    await _sembrar(cliente, datos)

    respuestas = await asyncio.gather(
        cliente.get("/api/v1/reportes/export.xlsx", params=MES),
        cliente.get("/api/v1/reportes/export.pdf", params=MES),
    )
    codigos = sorted(r.status_code for r in respuestas)
    assert codigos == [200, 409]

    rechazada = next(r for r in respuestas if r.status_code == 409)
    assert rechazada.json()["error"] == "export_en_curso"


async def test_la_reserva_se_libera_al_terminar(cliente, datos):
    """Si el turno no se soltara, el usuario quedaría bloqueado para siempre."""
    await _sembrar(cliente, datos)

    for _ in range(3):
        r = await cliente.get("/api/v1/reportes/export.xlsx", params=MES)
        assert r.status_code == 200


async def test_la_reserva_se_libera_aunque_el_export_falle(cliente, datos):
    from app.services import export

    como(cliente, AUTH_UID_A)
    # Un rango inválido corta antes de reservar; uno válido con el armado roto, no
    original = export.excel
    export.excel = lambda *_: (_ for _ in ()).throw(RuntimeError("boom"))
    try:
        with pytest.raises(RuntimeError):
            await cliente.get("/api/v1/reportes/export.xlsx", params=MES)
    finally:
        export.excel = original

    assert (await cliente.get("/api/v1/reportes/export.xlsx", params=MES)).status_code == 200


async def test_dos_usuarios_distintos_no_se_bloquean(cliente, datos):
    """El límite es por usuario; el semáforo global solo los serializa."""
    import asyncio

    from tests.conftest import token_para

    await _sembrar(cliente, datos)

    async def pedir(uid: str):
        return await cliente.get(
            "/api/v1/reportes/export.xlsx",
            params=MES,
            cookies={"sb_access": token_para(uid)},
        )

    a, b = await asyncio.gather(pedir(AUTH_UID_A), pedir(AUTH_UID_B))
    assert a.status_code == 200
    assert b.status_code == 200
