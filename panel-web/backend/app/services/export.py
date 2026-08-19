"""Armado de los archivos de exportación: Excel y PDF.

Se construyen en memoria y se mandan enteros. Con el tope de filas del reporte los
archivos quedan en pocos cientos de kB, y streaming de verdad no valdría la
complejidad; lo que sí importa con 350 MB de RAM es que no se armen dos a la vez.
"""
import asyncio
import contextlib
from datetime import date, datetime
from io import BytesIO

#: Un export por vez en todo el proceso. Protege la memoria del servidor frente a
#: dos usuarios distintos: el segundo espera su turno o recibe 429.
_turno = asyncio.Semaphore(1)

#: Un export por usuario. Es otra cosa: acá no se hace cola, se rechaza. Pedir dos
#: archivos a la vez nunca es intencional — son dos pestañas o un doble clic — y
#: hacerlo esperar solo consigue que el usuario piense que se colgó.
_generando: set[int] = set()


class ExportEnCurso(Exception):
    """El usuario ya tiene un archivo armándose."""


@contextlib.contextmanager
def reserva(usuario_id: int):
    """Toma el turno del usuario o falla en el acto.

    El chequeo y el alta pasan sin `await` en medio, así que en un event loop de un
    solo hilo son atómicos: no hay ventana para que dos peticiones pasen juntas.
    """
    if usuario_id in _generando:
        raise ExportEnCurso
    _generando.add(usuario_id)
    try:
        yield
    finally:
        # También corre si el cliente aborta la descarga a mitad de camino
        _generando.discard(usuario_id)

MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]

ETIQUETA_GRUPO = {"categoria": "Categoría", "mes": "Mes", "cuenta": "Cuenta"}


def turno():
    """Context manager async para serializar los exports."""
    return _turno


def nombre_archivo(desde: date, hasta: date, extension: str) -> str:
    return f"finanzas_{desde:%Y%m%d}_{hasta:%Y%m%d}.{extension}"


def _titulo_periodo(desde: date, hasta: date) -> str:
    if desde.year == hasta.year and desde.month == hasta.month:
        return f"{MESES_ES[desde.month - 1]} {desde.year}"
    return f"{desde:%d/%m/%Y} — {hasta:%d/%m/%Y}"


def _fecha(v) -> str:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v or "")


def _clave_legible(clave: str, group_by: str) -> str:
    """Las claves de mes vienen como 'YYYY-MM' de la base."""
    if group_by == "mes" and clave and "-" in clave:
        anio, mes = clave.split("-")[:2]
        return f"{MESES_ES[int(mes) - 1]} {anio}"
    return clave or "—"


# ── Excel ────────────────────────────────────────────────────────────────────

def excel(reporte: dict, movimientos: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="5B4FE8")  # índigo del panel
    plata = '"S/ "#,##0.00'

    group_by = reporte["group_by"]
    desde, hasta = reporte["periodo"]["desde"], reporte["periodo"]["hasta"]

    # Hoja 1 · Resumen
    hoja = wb.active
    hoja.title = "Resumen"
    hoja["A1"] = f"Finanzas · {_titulo_periodo(desde, hasta)}"
    hoja["A1"].font = Font(bold=True, size=14)
    hoja["A2"] = f"Agrupado por {ETIQUETA_GRUPO.get(group_by, group_by).lower()}"
    hoja["A2"].font = Font(color="6B7280")

    cabeceras = [ETIQUETA_GRUPO.get(group_by, group_by), "Ingresos", "Gastos", "Neto", "Movimientos"]
    hoja.append([])
    hoja.append(cabeceras)
    for celda in hoja[4]:
        celda.font = encabezado
        celda.fill = fondo
        celda.alignment = Alignment(horizontal="center")

    for f in reporte["filas"]:
        hoja.append([
            _clave_legible(f["clave"], group_by),
            f["ingresos"], f["gastos"], f["neto"], f["n"],
        ])

    t = reporte["totales"]
    hoja.append([])
    hoja.append(["TOTAL", t["ingresos"], t["gastos"], t["neto"], t["n"]])
    for celda in hoja[hoja.max_row]:
        celda.font = Font(bold=True)

    for fila in hoja.iter_rows(min_row=5, min_col=2, max_col=4):
        for celda in fila:
            celda.number_format = plata

    hoja.freeze_panes = "A5"
    for i, ancho in enumerate((28, 14, 14, 14, 13), start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho

    # Hoja 2 · Detalle
    det = wb.create_sheet("Movimientos")
    det.append(["Fecha", "Tipo", "Categoría", "Descripción", "Cuenta", "Monto"])
    for celda in det[1]:
        celda.font = encabezado
        celda.fill = fondo
    for m in movimientos:
        det.append([
            _fecha(m["fecha"]),
            "Ingreso" if m["tipo"] == "ingreso" else "Gasto",
            m["categoria"] or "",
            m["descripcion"] or "",
            m["cuenta"] or "",
            m["monto"] if m["tipo"] == "ingreso" else -m["monto"],
        ])
    for fila in det.iter_rows(min_row=2, min_col=6, max_col=6):
        for celda in fila:
            celda.number_format = plata
    det.freeze_panes = "A2"
    for i, ancho in enumerate((12, 10, 20, 38, 16, 14), start=1):
        det.column_dimensions[get_column_letter(i)].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────

def pdf(reporte: dict, movimientos: list[dict]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    group_by = reporte["group_by"]
    desde, hasta = reporte["periodo"]["desde"], reporte["periodo"]["hasta"]
    indigo = colors.HexColor("#5b4fe8")
    tinta2 = colors.HexColor("#6b7280")
    linea = colors.HexColor("#e5e7eb")

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle("titulo", parent=estilos["Title"], fontSize=17,
                            alignment=0, spaceAfter=2, textColor=colors.HexColor("#111827"))
    bajada = ParagraphStyle("bajada", parent=estilos["Normal"], fontSize=9.5,
                            textColor=tinta2, spaceAfter=14)
    seccion = ParagraphStyle("seccion", parent=estilos["Heading2"], fontSize=12,
                             spaceBefore=16, spaceAfter=6)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"Finanzas {_titulo_periodo(desde, hasta)}", author="FinanzasBot",
    )

    def plata(v: float) -> str:
        return f"S/ {v:,.2f}"

    filtros = reporte.get("filtros") or {}
    activos = [f"{k}: {v}" for k, v in filtros.items() if v]
    piezas = [
        Paragraph(f"Finanzas · {_titulo_periodo(desde, hasta)}", titulo),
        Paragraph(
            f"Agrupado por {ETIQUETA_GRUPO.get(group_by, group_by).lower()}"
            + (f" · {' · '.join(activos)}" if activos else "")
            + f" · generado el {date.today():%d/%m/%Y}",
            bajada,
        ),
    ]

    t = reporte["totales"]
    resumen = Table(
        [["Ingresos", "Gastos", "Neto", "Movimientos"],
         [plata(t["ingresos"]), plata(t["gastos"]), plata(t["neto"]), str(t["n"])]],
        colWidths=[43 * mm] * 4,
    )
    resumen.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, 0), 8.5),
        ("TEXTCOLOR", (0, 0), (-1, 0), tinta2),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 13),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
        ("TOPPADDING", (0, 1), (-1, 1), 0),
        ("LINEBELOW", (0, 1), (-1, 1), 1, linea),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 10),
    ]))
    piezas += [resumen, Paragraph(ETIQUETA_GRUPO.get(group_by, group_by), seccion)]

    filas = [[ETIQUETA_GRUPO.get(group_by, group_by), "Ingresos", "Gastos", "Neto", "Mov."]]
    filas += [
        [_clave_legible(f["clave"], group_by), plata(f["ingresos"]), plata(f["gastos"]),
         plata(f["neto"]), str(f["n"])]
        for f in reporte["filas"]
    ]
    piezas.append(_tabla(Table, TableStyle, colors, filas, [58, 30, 30, 30, 16], indigo, linea, mm))

    if movimientos:
        piezas.append(Paragraph(f"Movimientos ({len(movimientos)})", seccion))
        det = [["Fecha", "Categoría", "Descripción", "Cuenta", "Monto"]]
        for m in movimientos:
            det.append([
                _fecha(m["fecha"]),
                (m["categoria"] or "")[:18],
                (m["descripcion"] or "")[:34],
                (m["cuenta"] or "")[:14],
                ("+" if m["tipo"] == "ingreso" else "−") + plata(m["monto"]),
            ])
        piezas.append(_tabla(Table, TableStyle, colors, det, [22, 30, 60, 26, 26], indigo, linea, mm))

    piezas.append(Spacer(1, 8 * mm))
    doc.build(piezas)
    return buffer.getvalue()


def _tabla(Table, TableStyle, colors, filas, anchos_mm, indigo, linea, mm):
    tabla = Table(filas, colWidths=[a * mm for a in anchos_mm], repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), indigo),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7f9")]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, linea),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]))
    return tabla
