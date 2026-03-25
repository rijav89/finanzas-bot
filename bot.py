"""
bot.py — FinanzasBot v3.0
- Registro de ingresos y saldo disponible
- Editar / eliminar transacciones
- Recordatorios de pagos fijos
"""

import os
import io
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.error import BadRequest
from telegram.ext import (
    ApplicationBuilder,
    MessageHandler,
    CommandHandler,
    ConversationHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
    JobQueue,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from config import TOKEN
from db import (
    obtener_o_crear_usuario,
    guardar_transaccion,
    obtener_total_mes,
    obtener_historial,
    obtener_resumen_categorias,
    obtener_transacciones_mes,
    guardar_ingreso,
    obtener_total_ingresos_mes,
    obtener_historial_ingresos,
    obtener_ultimas_transacciones,
    eliminar_transaccion,
    editar_transaccion,
    guardar_pago_fijo,
    obtener_pagos_fijos,
    eliminar_pago_fijo,
    obtener_pagos_fijos_del_dia,
    actualizar_medio_ultimas,
    actualizar_medio_transaccion,
    actualizar_medio_ingreso_reciente,
)
from ocr import procesar_voucher
from categorias import clasificar_gasto
from gastos_manual import detectar_intencion, extraer_gastos, extraer_ingreso, extraer_edicion
from graficos import generar_grafico_categorias, generar_grafico_resumen

# Estados de conversación
ESPERANDO_DESCRIPCION       = 1
REGISTRO_MANUAL_DESCRIPCION = 2
REGISTRO_MANUAL_MEDIO       = 3
INGRESO_MONTO               = 4
INGRESO_DESCRIPCION         = 5
PAGO_FIJO_DESC              = 6
PAGO_FIJO_MONTO             = 7
PAGO_FIJO_DIA               = 8
EDITAR_CAMPO                = 9

datos_pendientes         = {}
registro_manual_pendiente = {}
ingreso_pendiente        = {}
pago_fijo_pendiente      = {}
edicion_pendiente        = {}

EMOJIS_CATEGORIA = {
    "Comida":          "🍽️",
    "Supermercado":    "🛒",
    "Transporte":      "🚗",
    "Servicios":       "💡",
    "Salud":           "🏥",
    "Educacion":       "📚",
    "Ropa":            "👕",
    "Entretenimiento": "🎬",
    "Tecnologia":      "💻",
    "Finanzas":        "🏦",
    "Mascotas":        "🐾",
    "Belleza":         "💅",
    "Hogar":           "🏠",
    "Otros":           "📦",
}

RESPUESTA_FUERA_DE_TEMA = (
    "🤖 Hola! Soy *FinanzasBot*, tu asistente de finanzas personales.\n\n"
    "Estoy diseñado exclusivamente para ayudarte a registrar y analizar tus finanzas. Puedo:\n\n"
    "📸 Leer vouchers de *Yape* y *Plin* automáticamente\n"
    "✍️ Registrar gastos que me escribas en texto libre\n"
    "💵 Registrar tus *ingresos* (sueldo, freelance, ventas, etc.)\n"
    "💰 Mostrarte tu *saldo disponible* del mes\n"
    "📊 Resúmenes y estadísticas de tus gastos\n"
    "🔔 Recordarte tus *pagos fijos* mensuales\n"
    "📥 Exportar tus gastos a Excel\n\n"
    "Por ejemplo puedes escribirme:\n"
    "_\"Hoy gasté 50 soles en almuerzo y 20 en taxi\"_\n"
    "_\"Me pagaron 4000 de sueldo\"_\n"
    "_\"Quiero registrar un ingreso\"_\n\n"
    "¿Qué deseas hacer?"
)


def menu_principal() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("📊 Resumen", callback_data="resumen"),
            InlineKeyboardButton("🏷️ Categorías", callback_data="categorias"),
        ],
        [
            InlineKeyboardButton("📜 Historial", callback_data="historial"),
            InlineKeyboardButton("📥 Exportar Excel", callback_data="exportar"),
        ],
        [
            InlineKeyboardButton("💵 Ingresos", callback_data="ver_ingresos"),
            InlineKeyboardButton("💰 Saldo", callback_data="ver_saldo"),
        ],
        [
            InlineKeyboardButton("🔔 Pagos fijos", callback_data="ver_pagos"),
            InlineKeyboardButton("✏️ Editar gastos", callback_data="ver_editar"),
        ],
        [
            InlineKeyboardButton("❓ Ayuda", callback_data="ayuda"),
        ],
    ])


# ── /start ──────────────────────────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 *Bienvenido a FinanzasBot!*\n\n"
        "Envíame una foto de tu voucher de *Yape* o *Plin*, o escríbeme tus gastos directamente.\n\n"
        "_Ej: \"Hoy gasté 50 soles en almuerzo y 30 en taxi\"_\n\n"
        "¿Qué deseas hacer?",
        parse_mode="Markdown",
        reply_markup=menu_principal()
    )


# ── Manejar texto libre ─────────────────────────────────────────────────────

async def handle_texto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mensaje = update.message.text.strip()
    telegram_id = update.message.from_user.id
    usuario_id = obtener_o_crear_usuario(telegram_id)

    intencion = detectar_intencion(mensaje)

    if intencion == "FUERA_DE_TEMA":
        await update.message.reply_text(
            RESPUESTA_FUERA_DE_TEMA, parse_mode="Markdown", reply_markup=menu_principal()
        )

    elif intencion == "VER_RESUMEN":
        texto = await mostrar_resumen(usuario_id)
        await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=menu_principal())

    elif intencion == "VER_CATEGORIAS":
        texto = await mostrar_categorias(usuario_id)
        await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=menu_principal())

    elif intencion == "VER_HISTORIAL":
        texto, teclado = await mostrar_historial(usuario_id)
        if teclado:
            await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=teclado)
        else:
            await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=menu_principal())

    elif intencion == "EXPORTAR":
        buffer, total = await generar_excel(usuario_id)
        if not buffer:
            await update.message.reply_text("📭 No tienes transacciones este mes.", reply_markup=menu_principal())
            return
        mes_actual = datetime.now().strftime("%B %Y")
        nombre_archivo = f"finanzas_{datetime.now().strftime('%Y_%m')}.xlsx"
        await update.message.reply_document(
            document=buffer, filename=nombre_archivo,
            caption=f"📊 Reporte de {mes_actual} — Total: S/ {total:.2f}",
            reply_markup=menu_principal()
        )

    elif intencion == "AYUDA":
        await update.message.reply_text(
            "🤖 *FinanzasBot — Ayuda*\n\n"
            "📸 Envía una foto de tu voucher de Yape o Plin\n"
            "✍️ O escribe tus gastos directamente:\n"
            "_\"Hoy gasté 50 en almuerzo y 20 en taxi\"_\n\n"
            "También puedes pedirme en lenguaje natural:\n"
            "• _\"cuánto gasté este mes\"_ → resumen\n"
            "• _\"en qué gasté más\"_ → categorías\n"
            "• _\"mis últimos gastos\"_ → historial\n"
            "• _\"exportar mis gastos\"_ → Excel\n\n"
            "/cancelar — Cancelar registro en curso",
            parse_mode="Markdown", reply_markup=menu_principal()
        )

    elif intencion == "VER_SALDO":
        ingresos = obtener_total_ingresos_mes(usuario_id)
        gastos = float(obtener_total_mes(usuario_id))
        saldo = ingresos - gastos
        emoji_saldo = "✅" if saldo >= 0 else "🔴"
        porcentaje = (gastos / ingresos * 100) if ingresos > 0 else 0
        barra = "█" * min(int(porcentaje / 10), 10) + "░" * max(0, 10 - int(porcentaje / 10))
        mes = datetime.now().strftime("%B %Y")
        await update.message.reply_text(
            f"💰 *Balance de {mes}*\n"
            f"─────────────────────\n"
            f"📈 Ingresos:  S/ {ingresos:.2f}\n"
            f"📉 Gastos:    S/ {gastos:.2f}\n"
            f"─────────────────────\n"
            f"{emoji_saldo} Saldo:     S/ {saldo:.2f}\n\n"
            f"Usaste el {porcentaje:.1f}% de tus ingresos\n{barra}",
            parse_mode="Markdown", reply_markup=menu_principal()
        )

    elif intencion == "REGISTRAR_INGRESO":
        datos = extraer_ingreso(mensaje)
        monto = datos.get("monto", 0)
        descripcion = datos.get("descripcion", "Ingreso")
        medio = datos.get("medio", "No especificado")

        if monto and float(monto) > 0:
            # Tiene monto — preguntar medio si no fue detectado
            guardar_ingreso(usuario_id, monto, descripcion)
            if medio == "No especificado":
                teclado_medio = InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton("📱 Yape", callback_data="ing_medio_Yape"),
                        InlineKeyboardButton("💙 Plin", callback_data="ing_medio_Plin"),
                    ],
                    [
                        InlineKeyboardButton("🏦 Transferencia", callback_data="ing_medio_Transferencia"),
                        InlineKeyboardButton("💵 Efectivo", callback_data="ing_medio_Efectivo"),
                    ],
                    [
                        InlineKeyboardButton("⏭️ Omitir", callback_data="ing_medio_Manual"),
                    ],
                ])
                await update.message.reply_text(
                    f"✅ *Ingreso registrado*\n\n"
                    f"💵 Monto: S/ {float(monto):.2f}\n"
                    f"📝 {descripcion}\n\n"
                    f"¿Por qué medio recibiste el pago?",
                    parse_mode="Markdown", reply_markup=teclado_medio
                )
            else:
                await update.message.reply_text(
                    f"✅ *Ingreso registrado*\n\n"
                    f"💵 Monto: S/ {float(monto):.2f}\n"
                    f"📝 {descripcion}\n"
                    f"📱 Medio: {medio}\n\n"
                    f"¿Qué deseas hacer ahora?",
                    parse_mode="Markdown", reply_markup=menu_principal()
                )
        else:
            # No tiene monto — iniciar flujo guiado
            ingreso_pendiente[telegram_id] = {}
            await update.message.reply_text(
                "💵 *Registrar ingreso*\n\n¿Cuánto recibiste? _(solo el número, ej: 1500)_",
                parse_mode="Markdown"
            )

    elif intencion == "INICIAR_REGISTRO":
        # El usuario quiere registrar pero no dio detalles — iniciar flujo
        registro_manual_pendiente[telegram_id] = {}
        await update.message.reply_text(
            "📝 ¡Claro! ¿En qué gastaste y cuánto fue el monto?\n\n"
            "_Ej: \"50 soles en almuerzo\" o puedes listar varios gastos a la vez\"_",
            parse_mode="Markdown"
        )

    elif intencion == "REGISTRAR_GASTOS":
        await _procesar_y_guardar_gastos(update, usuario_id, mensaje)


async def _procesar_y_guardar_gastos(update, usuario_id, mensaje):
    """Extrae gastos del mensaje, detecta fecha, pregunta medio y guarda."""
    from datetime import datetime
    await update.message.reply_text("⏳ Analizando tus gastos...")
    gastos, fecha_str = extraer_gastos(mensaje)

    if not gastos:
        await update.message.reply_text(
            "⚠️ No pude identificar gastos en tu mensaje.\n"
            "Intenta ser más específico:\n"
            "_\"Gasté 50 soles en almuerzo y 20 en taxi\"_",
            parse_mode="Markdown", reply_markup=menu_principal()
        )
        return

    # Convertir fecha_str a datetime
    try:
        fecha_dt = datetime.strptime(fecha_str, "%Y-%m-%d")
    except Exception:
        fecha_dt = None

    hoy = datetime.now().strftime("%Y-%m-%d")
    es_fecha_distinta = fecha_str and fecha_str != hoy

    resumen_texto = "✅ *Gastos registrados:*\n\n"
    total = 0
    for gasto in gastos:
        monto = gasto.get("monto", 0)
        descripcion = gasto.get("descripcion", "Sin descripción")
        categoria = gasto.get("categoria", "Otros")
        emoji = EMOJIS_CATEGORIA.get(categoria, "📦")
        guardar_transaccion(
            usuario_id, monto=monto, medio="Manual",
            descripcion=descripcion, categoria=categoria,
            destinatario="—", fecha_voucher="—",
            fecha=fecha_dt
        )
        resumen_texto += f"{emoji} S/ {monto:.2f} — {descripcion} _{categoria}_\n"
        total += float(monto)

    resumen_texto += f"\n💰 *Total registrado: S/ {total:.2f}*"
    if es_fecha_distinta:
        fecha_legible = fecha_dt.strftime("%d/%m/%Y") if fecha_dt else fecha_str
        resumen_texto += f"\n📅 Fecha registrada: *{fecha_legible}*"

    teclado_medio = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("📱 Yape", callback_data="medio_Yape"),
            InlineKeyboardButton("💙 Plin", callback_data="medio_Plin"),
        ],
        [
            InlineKeyboardButton("💵 Efectivo", callback_data="medio_Efectivo"),
            InlineKeyboardButton("💳 Tarjeta", callback_data="medio_Tarjeta"),
        ],
        [
            InlineKeyboardButton("🏦 Transferencia", callback_data="medio_Transferencia"),
            InlineKeyboardButton("⏭️ Omitir", callback_data="medio_Manual"),
        ],
    ])

    await update.message.reply_text(
        resumen_texto + "\n\n¿Con qué medio pagaste?",
        parse_mode="Markdown",
        reply_markup=teclado_medio
    )


# ── Recibir descripción después de INICIAR_REGISTRO ─────────────────────────

async def handle_descripcion_manual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Recibe el detalle del gasto cuando el usuario inició con 'registrar un gasto'."""
    telegram_id = update.message.from_user.id
    mensaje = update.message.text.strip()

    if telegram_id not in registro_manual_pendiente:
        return await handle_texto(update, context)

    del registro_manual_pendiente[telegram_id]
    usuario_id = obtener_o_crear_usuario(telegram_id)
    await _procesar_y_guardar_gastos(update, usuario_id, mensaje)


# ── Procesar imagen ─────────────────────────────────────────────────────────

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    photo = update.message.photo[-1]
    file = await photo.get_file()
    file_path = f"temp_{update.message.message_id}.jpg"
    await file.download_to_drive(file_path)

    try:
        await update.message.reply_text("⏳ Leyendo texto del voucher...")
        monto, medio, destinatario, fecha = procesar_voucher(file_path)

        if monto == "No detectado":
            await update.message.reply_text(
                "⚠️ No pude leer el monto del voucher.\n"
                "Asegúrate de que la imagen sea nítida.",
                reply_markup=menu_principal()
            )
            os.remove(file_path)
            return ConversationHandler.END

        descripcion = update.message.caption
        telegram_id = update.message.from_user.id

        if descripcion:
            await registrar_transaccion(update, telegram_id, monto, medio, destinatario, fecha, descripcion)
            os.remove(file_path)
            return ConversationHandler.END
        else:
            datos_pendientes[telegram_id] = {
                "monto": monto, "medio": medio,
                "destinatario": destinatario, "fecha": fecha,
            }
            await update.message.reply_text(
                f"📋 *Voucher leído:*\n"
                f"💰 Monto: S/ {monto}\n"
                f"📱 Medio: {medio}\n"
                f"👤 Destinatario: {destinatario}\n\n"
                f"¿A qué correspondió este gasto? Escribe una descripción breve.\n"
                f"_(Ej: almuerzo con amigos, uber al trabajo, consulta médica)_",
                parse_mode="Markdown"
            )
            os.remove(file_path)
            return ESPERANDO_DESCRIPCION

    except Exception as e:
        await update.message.reply_text(
            f"❌ Error al procesar la imagen: {e}",
            reply_markup=menu_principal()
        )
        if os.path.exists(file_path):
            os.remove(file_path)
        return ConversationHandler.END


# ── Recibir descripción pendiente ───────────────────────────────────────────

async def handle_descripcion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.message.from_user.id
    descripcion = update.message.text.strip()

    if telegram_id not in datos_pendientes:
        await update.message.reply_text(
            "⚠️ No encontré un voucher pendiente. Envía la imagen nuevamente.",
            reply_markup=menu_principal()
        )
        return ConversationHandler.END

    datos = datos_pendientes.pop(telegram_id)
    await registrar_transaccion(
        update, telegram_id,
        datos["monto"], datos["medio"], datos["destinatario"], datos["fecha"],
        descripcion
    )
    return ConversationHandler.END


# ── Registrar transacción desde voucher ─────────────────────────────────────

async def registrar_transaccion(update, telegram_id, monto, medio, destinatario, fecha, descripcion):
    try:
        categoria = clasificar_gasto(descripcion)
        emoji = EMOJIS_CATEGORIA.get(categoria, "📦")
        usuario_id = obtener_o_crear_usuario(telegram_id)
        guardar_transaccion(usuario_id, monto, medio, descripcion, categoria, destinatario, fecha)

        # Escapar caracteres especiales de Markdown en campos dinámicos
        def esc(text):
            for ch in ('_', '*', '`', '['):
                text = str(text).replace(ch, f'\\{ch}')
            return text

        await update.message.reply_text(
            f"✅ *Transacción registrada*\n\n"
            f"💰 Monto:         S/ {esc(monto)}\n"
            f"📱 Medio:          {esc(medio)}\n"
            f"👤 Destinatario:  {esc(destinatario)}\n"
            f"📅 Fecha:          {esc(fecha)}\n"
            f"📝 Descripción:   {esc(descripcion)}\n"
            f"{emoji} Categoría:    {esc(categoria)}\n\n"
            f"¿Qué deseas hacer ahora?",
            parse_mode="Markdown",
            reply_markup=menu_principal()
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Error al registrar: {e}", reply_markup=menu_principal())


# ── Lógica compartida de vistas ──────────────────────────────────────────────

async def mostrar_resumen(usuario_id):
    total = obtener_total_mes(usuario_id)
    return f"📊 *Total gastado este mes:* S/ {float(total):.2f}\n\nUsa *Categorías* para ver el desglose."


async def mostrar_categorias(usuario_id):
    filas = obtener_resumen_categorias(usuario_id)
    total = obtener_total_mes(usuario_id)
    if not filas:
        return "📭 No tienes transacciones este mes."

    def esc(text):
        for ch in ('_', '*', '`', '['):
            text = str(text).replace(ch, f'\\{ch}')
        return text

    mes_actual = datetime.now().strftime("%B %Y")
    texto = f"📊 *Resumen de {mes_actual}:*\n\n"
    for categoria, subtotal, cantidad in filas:
        emoji = EMOJIS_CATEGORIA.get(categoria, "📦")
        porcentaje = (float(subtotal) / float(total) * 100) if total else 0
        barra = "█" * int(porcentaje / 10) + "░" * (10 - int(porcentaje / 10))
        texto += (
            f"{emoji} *{esc(categoria)}*\n"
            f"   S/ {subtotal:.2f} · {cantidad} transac. · {porcentaje:.1f}%\n"
            f"   {barra}\n\n"
        )
    texto += f"💰 *Total: S/ {float(total):.2f}*"
    return texto


<<<<<<< HEAD
async def mostrar_historial(usuario_id):
    filas = obtener_historial(usuario_id, limite=10)
    if not filas:
        return "📭 No tienes transacciones registradas aún.", None

    texto = "📜 Últimas 10 transacciones:\n"
=======
async def mostrar_historial(usuario_id, page=0):
    limite = 5
    offset = page * limite
    filas = obtener_historial(usuario_id, limite=limite + 1, offset=offset)
    
    hay_mas = len(filas) > limite
    filas_mostrar = filas[:limite]

    if not filas_mostrar and page == 0:
        return "📭 No tienes transacciones registradas aún.", None
    elif not filas_mostrar:
        return "📭 No hay más transacciones.", None

    texto = f"📜 *Historial de transacciones (Pág {page + 1}):*\n"
>>>>>>> c46aaa2 (feat: Mejoras visuales (graficos matplotlib), paginacion y UX)
    texto += "─────────────────────\n\n"
    for i, (monto, medio, descripcion, categoria, destinatario, fecha) in enumerate(filas_mostrar, 1):
        emoji = EMOJIS_CATEGORIA.get(categoria, "📦")
        fecha_str = fecha.strftime("%d/%m/%Y  %H:%M") if fecha else "—"

        if destinatario and destinatario not in ("—", "No detectado"):
            partes = destinatario.strip().split()
            if len(partes) >= 4:
                dest_corto = f"{partes[0]} {partes[2]}"
            elif len(partes) == 3:
                dest_corto = f"{partes[0]} {partes[1]}"
            else:
                dest_corto = destinatario
        else:
            dest_corto = None

        texto += f"{emoji} {categoria}  ·  S/ {float(monto):.2f}\n"
        texto += f"📝 {descripcion}\n"
        texto += f"📱 {medio}\n"
        texto += f"📅 {fecha_str}\n"
        if dest_corto:
            texto += f"👤 {dest_corto}\n"
        texto += "─────────────────────\n"
<<<<<<< HEAD
    return texto, None
=======

    nav_botones = []
    if page > 0:
        nav_botones.append(InlineKeyboardButton("⬅️ Anterior", callback_data=f"historial_{page-1}"))
    if hay_mas:
        nav_botones.append(InlineKeyboardButton("Siguiente ➡️", callback_data=f"historial_{page+1}"))
    
    botones = []
    if nav_botones:
        botones.append(nav_botones)
    botones.append([InlineKeyboardButton("🔙 Menú Principal", callback_data="menu_principal")])
    
    return texto, InlineKeyboardMarkup(botones)
>>>>>>> c46aaa2 (feat: Mejoras visuales (graficos matplotlib), paginacion y UX)


async def generar_excel(usuario_id):
    filas = obtener_transacciones_mes(usuario_id)
    if not filas:
        return None, 0
    wb = openpyxl.Workbook()
    ws = wb.active
    mes_actual = datetime.now().strftime("%B %Y")
    ws.title = mes_actual
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["#", "Monto (S/)", "Medio", "Descripción", "Categoría", "Destinatario", "Fecha Voucher", "Fecha Registro"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    total = 0
    for i, (id_, monto, medio, descripcion, categoria, destinatario, fecha_voucher, fecha) in enumerate(filas, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=float(monto))
        ws.cell(row=i, column=3, value=medio)
        ws.cell(row=i, column=4, value=descripcion)
        ws.cell(row=i, column=5, value=categoria)
        ws.cell(row=i, column=6, value=destinatario)
        ws.cell(row=i, column=7, value=fecha_voucher)
        ws.cell(row=i, column=8, value=fecha.strftime("%d/%m/%Y %H:%M") if fecha else "")
        total += float(monto)
    fila_total = len(filas) + 2
    ws.cell(row=fila_total, column=1, value="TOTAL")
    ws.cell(row=fila_total, column=2, value=total).font = Font(bold=True)
    anchos = [5, 12, 10, 35, 18, 25, 18, 20]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, total


# ── Comandos ─────────────────────────────────────────────────────────────────

async def resumen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    usuario_id = obtener_o_crear_usuario(update.message.from_user.id)
    texto = await mostrar_resumen(usuario_id)
    await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=menu_principal())

async def categorias(update: Update, context: ContextTypes.DEFAULT_TYPE):
    usuario_id = obtener_o_crear_usuario(update.message.from_user.id)
    texto = await mostrar_categorias(usuario_id)
    await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=menu_principal())

async def historial(update: Update, context: ContextTypes.DEFAULT_TYPE):
    usuario_id = obtener_o_crear_usuario(update.message.from_user.id)
<<<<<<< HEAD
    texto, _ = await mostrar_historial(usuario_id)
    await update.message.reply_text(texto, reply_markup=menu_principal())
=======
    texto, teclado = await mostrar_historial(usuario_id)
    if teclado:
        await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=teclado)
    else:
        await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=menu_principal())
>>>>>>> c46aaa2 (feat: Mejoras visuales (graficos matplotlib), paginacion y UX)

async def exportar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    usuario_id = obtener_o_crear_usuario(update.message.from_user.id)
    buffer, total = await generar_excel(usuario_id)
    if not buffer:
        await update.message.reply_text("📭 No tienes transacciones este mes.", reply_markup=menu_principal())
        return
    mes_actual = datetime.now().strftime("%B %Y")
    nombre_archivo = f"finanzas_{datetime.now().strftime('%Y_%m')}.xlsx"
    await update.message.reply_document(
        document=buffer, filename=nombre_archivo,
        caption=f"📊 Reporte de {mes_actual} — Total: S/ {total:.2f}",
        reply_markup=menu_principal()
    )

async def ayuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🤖 *FinanzasBot — Ayuda*\n\n"
        "📸 Envía una foto de tu voucher de Yape o Plin\n"
        "✍️ O escribe tus gastos directamente:\n"
        "_\"Hoy gasté 50 en almuerzo y 20 en taxi\"_\n\n"
        "/resumen · /categorias · /historial · /exportar\n"
        "/cancelar — Cancelar registro en curso",
        parse_mode="Markdown",
        reply_markup=menu_principal()
    )

async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid = update.message.from_user.id
    datos_pendientes.pop(tid, None)
    ingreso_pendiente.pop(tid, None)
    pago_fijo_pendiente.pop(tid, None)
    edicion_pendiente.pop(tid, None)
    registro_manual_pendiente.pop(tid, None)
    await update.message.reply_text("❌ Operación cancelada.", reply_markup=menu_principal())
    return ConversationHandler.END


# ── /ingreso ─────────────────────────────────────────────────────────────────

async def cmd_ingreso(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid = update.message.from_user.id
    ingreso_pendiente[tid] = {}
    await update.message.reply_text(
        "💵 *Registrar ingreso*\n\n¿Cuánto recibiste? _(solo el número, ej: 1500)_\n\n/cancelar para salir.",
        parse_mode="Markdown"
    )
    return INGRESO_MONTO

async def ingreso_recibir_monto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid = update.message.from_user.id
    try:
        monto = float(update.message.text.strip().replace(",", "."))
        ingreso_pendiente[tid]["monto"] = monto
        await update.message.reply_text(
            f"✅ Monto: S/ {monto:.2f}\n\n¿De dónde proviene este ingreso?\n_(Ej: sueldo, freelance, venta, bono)_"
        )
        return INGRESO_DESCRIPCION
    except ValueError:
        await update.message.reply_text("⚠️ Ingresa solo el número. Ej: *1500* o *250.50*", parse_mode="Markdown")
        return INGRESO_MONTO

async def ingreso_recibir_descripcion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid = update.message.from_user.id
    descripcion = update.message.text.strip()
    datos = ingreso_pendiente.pop(tid, {})
    monto = datos.get("monto", 0)
    categoria = clasificar_gasto(descripcion) if descripcion else "Ingreso"
    usuario_id = obtener_o_crear_usuario(tid)
    guardar_ingreso(usuario_id, monto, descripcion, categoria)
    await update.message.reply_text(
        f"✅ *Ingreso registrado*\n\n"
        f"💵 Monto: S/ {monto:.2f}\n"
        f"📝 Descripción: {descripcion}\n\n"
        f"Usa /saldo para ver tu balance del mes.",
        parse_mode="Markdown",
        reply_markup=menu_principal()
    )
    return ConversationHandler.END


# ── /saldo ────────────────────────────────────────────────────────────────────

async def cmd_saldo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    usuario_id = obtener_o_crear_usuario(update.message.from_user.id)
    ingresos = obtener_total_ingresos_mes(usuario_id)
    gastos = float(obtener_total_mes(usuario_id))
    saldo = ingresos - gastos
    emoji_saldo = "✅" if saldo >= 0 else "🔴"
    porcentaje = (gastos / ingresos * 100) if ingresos > 0 else 0
    barra_usada = int(porcentaje / 10)
    barra = "█" * min(barra_usada, 10) + "░" * max(0, 10 - barra_usada)
    mes = datetime.now().strftime("%B %Y")
    await update.message.reply_text(
        f"💰 *Balance de {mes}*\n"
        f"─────────────────────\n"
        f"📈 Ingresos:   S/ {ingresos:.2f}\n"
        f"📉 Gastos:     S/ {gastos:.2f}\n"
        f"─────────────────────\n"
        f"{emoji_saldo} Saldo:      S/ {saldo:.2f}\n\n"
        f"Usaste el {porcentaje:.1f}% de tus ingresos\n"
        f"{barra}",
        parse_mode="Markdown",
        reply_markup=menu_principal()
    )


# ── /pagos ─────────────────────────────────────────────────────────────────────

async def cmd_pagos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    usuario_id = obtener_o_crear_usuario(update.message.from_user.id)
    pagos = obtener_pagos_fijos(usuario_id)
    hoy = datetime.now().day
    texto = "🔔 *Pagos fijos registrados:*\n─────────────────────\n\n"
    if pagos:
        for id_, desc, monto, dia, cat in pagos:
            dias_restantes = dia - hoy
            if dias_restantes < 0:
                dias_restantes += 30
            if dias_restantes == 0:
                estado = "⚠️ *¡Vence hoy!*"
            elif dias_restantes <= 3:
                estado = f"🔴 Vence en {dias_restantes} días"
            else:
                estado = f"📅 Día {dia} de cada mes"
            texto += f"• {desc} — S/ {monto:.2f}\n  {estado}\n\n"
    else:
        texto += "_No tienes pagos fijos registrados._\n\n"

    teclado = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Agregar pago fijo", callback_data="agregar_pago_fijo")],
        [InlineKeyboardButton("🗑️ Eliminar pago fijo", callback_data="listar_eliminar_pago")],
    ])
    await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=teclado)


# ── Flujo agregar pago fijo ───────────────────────────────────────────────────

async def pago_fijo_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid = update.message.from_user.id
    pago_fijo_pendiente[tid] = {"descripcion": update.message.text.strip()}
    await update.message.reply_text("💵 ¿Cuánto es el monto? _(ej: 99)_")
    return PAGO_FIJO_MONTO

async def pago_fijo_monto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid = update.message.from_user.id
    try:
        monto = float(update.message.text.strip().replace(",", "."))
        pago_fijo_pendiente[tid]["monto"] = monto
        await update.message.reply_text("📅 ¿Qué día del mes vence? _(del 1 al 31)_")
        return PAGO_FIJO_DIA
    except ValueError:
        await update.message.reply_text("⚠️ Ingresa solo el número. Ej: *99*", parse_mode="Markdown")
        return PAGO_FIJO_MONTO

async def pago_fijo_dia(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid = update.message.from_user.id
    try:
        dia = int(update.message.text.strip())
        if not 1 <= dia <= 31:
            raise ValueError
        datos = pago_fijo_pendiente.pop(tid, {})
        usuario_id = obtener_o_crear_usuario(tid)
        categoria = clasificar_gasto(datos["descripcion"])
        guardar_pago_fijo(usuario_id, datos["descripcion"], datos["monto"], dia, categoria)
        await update.message.reply_text(
            f"✅ *Pago fijo registrado*\n\n"
            f"📝 {datos['descripcion']}\n"
            f"💵 S/ {datos['monto']:.2f}\n"
            f"📅 Vence el día {dia} de cada mes\n\n"
            f"Te recordaré automáticamente ese día.",
            parse_mode="Markdown",
            reply_markup=menu_principal()
        )
        return ConversationHandler.END
    except ValueError:
        await update.message.reply_text("⚠️ Ingresa un número entre 1 y 31.")
        return PAGO_FIJO_DIA


# ── /editar ────────────────────────────────────────────────────────────────────

async def cmd_editar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    usuario_id = obtener_o_crear_usuario(update.message.from_user.id)
    # Llama al handler con página 0 directamente reutilizando lógica
    filas = obtener_ultimas_transacciones(usuario_id, limite=6, offset=0)
    hay_mas = len(filas) > 5
    filas_mostrar = filas[:5]
    if not filas_mostrar:
        await update.message.reply_text("📭 No tienes transacciones para editar.", reply_markup=menu_principal())
        return
    texto = "✏️ *Gestión de transacciones (Pág 1):*\n─────────────────────\n\n"
    botones = []
    for id_, monto, desc, cat, medio, fecha in filas_mostrar:
        emoji = EMOJIS_CATEGORIA.get(cat, "📦")
        fecha_str = fecha.strftime("%d/%m %H:%M") if fecha else "—"
        texto += f"`#{id_}` {emoji} S/ {float(monto):.2f} — {desc}\n📅 {fecha_str}\n\n"
        botones.append([
            InlineKeyboardButton(f"✏️ Editar #{id_}", callback_data=f"editar_{id_}"),
            InlineKeyboardButton(f"🗑️ Eliminar #{id_}", callback_data=f"eliminar_{id_}"),
        ])
    nav_botones = []
    if hay_mas:
        nav_botones.append(InlineKeyboardButton("Siguiente ➡️", callback_data="ver_editar_1"))
    if nav_botones:
        botones.append(nav_botones)
    botones.append([InlineKeyboardButton("🔙 Volver", callback_data="menu_principal")])
    await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))


# ── Recordatorio automático de pagos fijos ────────────────────────────────────

async def enviar_recordatorios(context: ContextTypes.DEFAULT_TYPE):
    hoy = datetime.now().day
    pagos = obtener_pagos_fijos_del_dia(hoy)
    for _, usuario_id, telegram_id, descripcion, monto, categoria in pagos:
        try:
            await context.bot.send_message(
                chat_id=telegram_id,
                text=(
                    f"🔔 *Recordatorio de pago fijo*\n\n"
                    f"📝 {descripcion}\n"
                    f"💵 S/ {float(monto):.2f}\n\n"
                    f"¿Ya lo pagaste? Envíame el voucher o regístralo manualmente."
                ),
                parse_mode="Markdown"
            )
        except Exception:
            pass


# ── Callback botones inline ──────────────────────────────────────────────────

async def safe_edit(query, texto, **kwargs):
    """Edita un mensaje ignorando el error si el contenido no cambió."""
    try:
        await query.edit_message_text(texto, **kwargs)
    except BadRequest as e:
        err_msg = str(e)
        if "Message is not modified" in err_msg:
            pass
        elif "There is no text in the message to edit" in err_msg:
            # Si el mensaje actual es una foto/documento, se borra y se envía el texto como nuevo
            await query.message.delete()
            await query.message.reply_text(texto, **kwargs)
        else:
            raise



async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    usuario_id = obtener_o_crear_usuario(query.from_user.id)
    accion = query.data

    # ── Medio de edición ──────────────────────────────────────────────
    if accion.startswith("edit_medio_"):
        partes = accion.replace("edit_medio_", "").split("_", 1)
        trans_id = int(partes[0])
        medio = partes[1]
        if medio != "skip":
            actualizar_medio_transaccion(trans_id, usuario_id, medio)
        nuevo_texto = query.message.text.replace("\n\n¿Actualizar el medio de pago?", f"\n📱 Medio: *{medio}*" if medio != "skip" else "")
        await safe_edit(query, nuevo_texto, parse_mode="Markdown", reply_markup=menu_principal())

    # ── Medio de ingreso ───────────────────────────────────────────────
    elif accion.startswith("ing_medio_"):
        medio = accion.replace("ing_medio_", "")
        actualizar_medio_ingreso_reciente(usuario_id, medio)
        await safe_edit(query,
            query.message.text.replace("\n\n¿Por qué medio recibiste el pago?", f"\n📱 Medio: *{medio}*"),
            parse_mode="Markdown", reply_markup=menu_principal()
        )

    # ── Selección de medio de pago (gastos) ───────────────────────────
    elif accion.startswith("medio_"):
        medio = accion.replace("medio_", "")
        actualizar_medio_ultimas(usuario_id, medio)
        await safe_edit(query,
            query.message.text.replace("\n\n¿Con qué medio pagaste?", f"\n\n📱 Medio: *{medio}*"),
            parse_mode="Markdown", reply_markup=menu_principal()
        )

    # ── Eliminar transacción ───────────────────────────────────────────
    elif accion.startswith("eliminar_"):
        trans_id = int(accion.replace("eliminar_", ""))
        botones = [
            [
                InlineKeyboardButton("✅ Sí, eliminar", callback_data=f"conf_elim_{trans_id}"),
                InlineKeyboardButton("❌ No, cancelar", callback_data="ver_editar")
            ]
        ]
        await safe_edit(query, f"⚠️ ¿Estás seguro de que deseas eliminar la transacción #{trans_id}?", reply_markup=InlineKeyboardMarkup(botones))

    elif accion.startswith("conf_elim_"):
        trans_id = int(accion.replace("conf_elim_", ""))
        eliminado = eliminar_transaccion(trans_id, usuario_id)
        if eliminado:
            await safe_edit(query, f"🗑️ Transacción #{trans_id} eliminada.", reply_markup=menu_principal())
        else:
            await safe_edit(query, "⚠️ No se pudo eliminar.", reply_markup=menu_principal())

    # ── Iniciar edición ────────────────────────────────────────────────
    elif accion.startswith("editar_"):
        trans_id = int(accion.replace("editar_", ""))
        edicion_pendiente[query.from_user.id] = {"id": trans_id}
        await query.message.reply_text(
            f"✏️ Editando transacción *#{trans_id}*\n\n"
            f"Escribe los nuevos datos como quieras, por ejemplo:\n"
            f"_\"50 comida de gato\"_\n"
            f"_\"35.50 almuerzo con el equipo\"_\n"
            f"_\"120 polo de trabajo\"_\n\n"
            f"/cancelar para salir.",
            parse_mode="Markdown"
        )

    # ── Agregar pago fijo desde botón ──────────────────────────────────
    elif accion == "agregar_pago_fijo":
        pago_fijo_pendiente[query.from_user.id] = {}
        await query.message.reply_text(
            "📝 ¿Cómo se llama este pago fijo?\n_(Ej: Internet Movistar, Alquiler, Netflix)_\n\n/cancelar para salir."
        )

    # ── Listar para eliminar pago fijo ─────────────────────────────────
    elif accion == "listar_eliminar_pago":
        pagos = obtener_pagos_fijos(usuario_id)
        if not pagos:
            await safe_edit(query, "📭 No tienes pagos fijos registrados.", reply_markup=menu_principal())
            return
        botones = [[InlineKeyboardButton(f"🗑️ {desc} — S/{monto}", callback_data=f"del_pago_{id_}")] for id_, desc, monto, dia, cat in pagos]
        botones.append([InlineKeyboardButton("🔙 Volver", callback_data="ver_pagos")])
        await safe_edit(query, "Selecciona el pago a eliminar:", reply_markup=InlineKeyboardMarkup(botones))

    elif accion.startswith("del_pago_"):
        pago_id = int(accion.replace("del_pago_", ""))
        eliminar_pago_fijo(pago_id, usuario_id)
        await safe_edit(query, "✅ Pago fijo eliminado.", reply_markup=menu_principal())

    elif accion == "ver_ingresos":
        filas = obtener_historial_ingresos(usuario_id, limite=5)
        total = obtener_total_ingresos_mes(usuario_id)
        if not filas:
            await safe_edit(query, "📭 No tienes ingresos registrados este mes.\n\nUsa /ingreso para agregar uno.", reply_markup=menu_principal())
            return
        texto = "💵 *Ingresos del mes:*\n─────────────────────\n\n"
        for monto, desc, cat, fecha in filas:
            fecha_str = fecha.strftime("%d/%m/%Y") if fecha else "—"
            texto += f"💵 S/ {float(monto):.2f} — {desc}\n📅 {fecha_str}\n\n"
        texto += f"─────────────────────\n💵 *Total: S/ {total:.2f}*"
        await safe_edit(query, texto, parse_mode="Markdown", reply_markup=menu_principal())

    elif accion == "ver_saldo":
        ingresos = obtener_total_ingresos_mes(usuario_id)
        gastos = float(obtener_total_mes(usuario_id))
        saldo = ingresos - gastos
        emoji_saldo = "✅" if saldo >= 0 else "🔴"
        porcentaje = (gastos / ingresos * 100) if ingresos > 0 else 0
        barra_usada = int(porcentaje / 10)
        barra = "█" * min(barra_usada, 10) + "░" * max(0, 10 - barra_usada)
        mes = datetime.now().strftime("%B %Y")
        await safe_edit(query,
            f"💰 *Balance de {mes}*\n"
            f"─────────────────────\n"
            f"📈 Ingresos:  S/ {ingresos:.2f}\n"
            f"📉 Gastos:    S/ {gastos:.2f}\n"
            f"─────────────────────\n"
            f"{emoji_saldo} Saldo:     S/ {saldo:.2f}\n\n"
            f"Usaste el {porcentaje:.1f}% de tus ingresos\n{barra}",
            parse_mode="Markdown", reply_markup=menu_principal()
        )

    elif accion == "ver_pagos":
        pagos = obtener_pagos_fijos(usuario_id)
        hoy = datetime.now().day
        texto = "🔔 *Pagos fijos:*\n─────────────────────\n\n"
        if pagos:
            for id_, desc, monto, dia, cat in pagos:
                dias_restantes = dia - hoy
                if dias_restantes < 0:
                    dias_restantes += 30
                estado = "⚠️ *¡Vence hoy!*" if dias_restantes == 0 else (f"🔴 En {dias_restantes} días" if dias_restantes <= 3 else f"📅 Día {dia}")
                texto += f"• {desc} — S/ {monto:.2f}\n  {estado}\n\n"
        else:
            texto += "_No tienes pagos fijos._\n\n"
        teclado = InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ Agregar", callback_data="agregar_pago_fijo")],
            [InlineKeyboardButton("🗑️ Eliminar", callback_data="listar_eliminar_pago")],
            [InlineKeyboardButton("🔙 Volver", callback_data="menu_principal")],
        ])
        await safe_edit(query, texto, parse_mode="Markdown", reply_markup=teclado)

    elif accion.startswith("ver_editar"):
        page = 0
        if "_" in accion and accion != "ver_editar":
            partes = accion.split("_")
            if len(partes) > 2 and partes[2].isdigit():
                page = int(partes[2])
                
        limite = 5
        offset = page * limite
        filas = obtener_ultimas_transacciones(usuario_id, limite=limite+1, offset=offset)
        hay_mas = len(filas) > limite
        filas_mostrar = filas[:limite]
        
        if not filas_mostrar and page == 0:
            await query.answer("📭 No tienes transacciones para gestionar.")
            return
        elif not filas_mostrar:
            await query.answer("📭 No hay más transacciones.")
            return
            
        texto = f"✏️ *Gestionar transacciones (Pág {page+1}):*\n─────────────────────\n\n"
        botones = []
        for id_, monto, desc, cat, medio, fecha in filas_mostrar:
            emoji = EMOJIS_CATEGORIA.get(cat, "📦")
            fecha_str = fecha.strftime("%d/%m %H:%M") if fecha else "—"
            texto += f"`#{id_}` {emoji} S/ {float(monto):.2f} — {desc}\n📅 {fecha_str}\n\n"
            botones.append([
                InlineKeyboardButton(f"✏️ Editar #{id_}", callback_data=f"editar_{id_}"),
                InlineKeyboardButton(f"🗑️ Eliminar #{id_}", callback_data=f"eliminar_{id_}"),
            ])
            
        nav_botones = []
        if page > 0:
            nav_botones.append(InlineKeyboardButton("⬅️ Anterior", callback_data=f"ver_editar_{page-1}"))
        if hay_mas:
            nav_botones.append(InlineKeyboardButton("Siguiente ➡️", callback_data=f"ver_editar_{page+1}"))
            
        if nav_botones:
            botones.append(nav_botones)
        botones.append([InlineKeyboardButton("🔙 Volver", callback_data="menu_principal")])
        await safe_edit(query, texto, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))
        
    elif accion == "menu_principal":
        await safe_edit(query,
            "👋 *Bienvenido a FinanzasBot!*\n\n"
            "Envíame una foto de tu voucher de *Yape* o *Plin*, o escríbeme tus gastos directamente.\n\n"
            "_Ej: \"Hoy gasté 50 soles en almuerzo y 30 en taxi\"_\n\n"
            "¿Qué deseas hacer?",
            parse_mode="Markdown",
            reply_markup=menu_principal()
        )

    elif accion == "resumen":
        texto = await mostrar_resumen(usuario_id)
        ingresos = obtener_total_ingresos_mes(usuario_id)
        gastos = float(obtener_total_mes(usuario_id))
        mes = datetime.now().strftime("%B %Y")
        img_buffer = generar_grafico_resumen(ingresos, gastos, mes)
        
        if query.message.photo or query.message.document:
            await query.message.delete()
            
        if img_buffer:
            await context.bot.send_photo(chat_id=query.message.chat_id, photo=img_buffer, caption=texto, parse_mode="Markdown", reply_markup=menu_principal())
        else:
            await context.bot.send_message(chat_id=query.message.chat_id, text=texto, parse_mode="Markdown", reply_markup=menu_principal())
            
    elif accion == "categorias":
        texto = await mostrar_categorias(usuario_id)
<<<<<<< HEAD
        await safe_edit(query, texto, parse_mode="Markdown", reply_markup=menu_principal())
    elif accion == "historial":
        texto, _ = await mostrar_historial(usuario_id)
        await safe_edit(query, texto, reply_markup=menu_principal())
=======
        datos = obtener_resumen_categorias(usuario_id)
        total = obtener_total_mes(usuario_id)
        mes = datetime.now().strftime("%B %Y")
        img_buffer = generar_grafico_categorias(datos, float(total or 0), mes)
        
        if query.message.photo or query.message.document:
            await query.message.delete()
            
        if img_buffer:
            await context.bot.send_photo(chat_id=query.message.chat_id, photo=img_buffer, caption=texto, parse_mode="Markdown", reply_markup=menu_principal())
        else:
            await context.bot.send_message(chat_id=query.message.chat_id, text=texto, parse_mode="Markdown", reply_markup=menu_principal())
            
    elif accion.startswith("historial"):
        page = 0
        if "_" in accion and accion != "historial":
            partes = accion.split("_")
            if len(partes) > 1 and partes[1].isdigit():
                page = int(partes[1])
                
        texto, teclado = await mostrar_historial(usuario_id, page)
        
        if teclado:
            await safe_edit(query, texto, parse_mode="Markdown", reply_markup=teclado)
        else:
            await safe_edit(query, texto, parse_mode="Markdown", reply_markup=menu_principal())
        
>>>>>>> c46aaa2 (feat: Mejoras visuales (graficos matplotlib), paginacion y UX)
    elif accion == "exportar":
        buffer, total = await generar_excel(usuario_id)
        if not buffer:
            await safe_edit(query, "📭 No tienes transacciones este mes.", reply_markup=menu_principal())
            return
        mes_actual = datetime.now().strftime("%B %Y")
        nombre_archivo = f"finanzas_{datetime.now().strftime('%Y_%m')}.xlsx"
        await query.message.delete()
        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=buffer, filename=nombre_archivo,
            caption=f"📊 Reporte de {mes_actual} — Total: S/ {total:.2f}",
            reply_markup=menu_principal()
        )
    elif accion == "ayuda":
        await safe_edit(query,
            "🤖 *FinanzasBot — Ayuda*\n\n"
            "📸 Envía una foto de tu voucher de Yape o Plin\n"
            "✍️ O escribe tus gastos directamente:\n"
            "_\"Hoy gasté 50 en almuerzo y 20 en taxi\"_\n\n"
            "/resumen · /categorias · /historial · /exportar\n"
            "/cancelar — Cancelar registro en curso",
            parse_mode="Markdown",
            reply_markup=menu_principal()
        )


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    app = ApplicationBuilder().token(TOKEN).build()

    # Voucher foto
    conv_voucher = ConversationHandler(
        entry_points=[MessageHandler(filters.PHOTO, handle_photo)],
        states={ESPERANDO_DESCRIPCION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_descripcion)]},
        fallbacks=[CommandHandler("cancelar", cancelar)],
    )

    # Registro de ingreso
    conv_ingreso = ConversationHandler(
        entry_points=[CommandHandler("ingreso", cmd_ingreso)],
        states={
            INGRESO_MONTO:       [MessageHandler(filters.TEXT & ~filters.COMMAND, ingreso_recibir_monto)],
            INGRESO_DESCRIPCION: [MessageHandler(filters.TEXT & ~filters.COMMAND, ingreso_recibir_descripcion)],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)],
    )

    # Pago fijo (cuando se inicia desde texto, no botón)
    conv_pago_fijo = ConversationHandler(
        entry_points=[CommandHandler("pagofijo", cmd_pagos)],
        states={
            PAGO_FIJO_DESC:  [MessageHandler(filters.TEXT & ~filters.COMMAND, pago_fijo_desc)],
            PAGO_FIJO_MONTO: [MessageHandler(filters.TEXT & ~filters.COMMAND, pago_fijo_monto)],
            PAGO_FIJO_DIA:   [MessageHandler(filters.TEXT & ~filters.COMMAND, pago_fijo_dia)],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)],
    )

    app.add_handler(conv_voucher)
    app.add_handler(conv_ingreso)
    app.add_handler(conv_pago_fijo)
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("resumen", resumen))
    app.add_handler(CommandHandler("categorias", categorias))
    app.add_handler(CommandHandler("historial", historial))
    app.add_handler(CommandHandler("exportar", exportar))
    app.add_handler(CommandHandler("saldo", cmd_saldo))
    app.add_handler(CommandHandler("pagos", cmd_pagos))
    app.add_handler(CommandHandler("editar", cmd_editar))
    app.add_handler(CommandHandler("ayuda", ayuda))
    app.add_handler(CommandHandler("cancelar", cancelar))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_texto_router))

    # Recordatorio diario a las 9am
    app.job_queue.run_daily(enviar_recordatorios, time=datetime.strptime("09:00", "%H:%M").time())

    print("Bot corriendo...")
    try:
        app.run_polling()
    except KeyboardInterrupt:
        pass


async def handle_texto_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tid = update.message.from_user.id

    # 1. Ingreso pendiente (monto o descripción)
    if tid in ingreso_pendiente:
        datos = ingreso_pendiente[tid]
        if "monto" not in datos:
            await ingreso_recibir_monto(update, context)
        else:
            await ingreso_recibir_descripcion(update, context)
        return

    # 2. Edición pendiente
    if tid in edicion_pendiente:
        trans_id = edicion_pendiente.pop(tid)["id"]
        try:
            datos = extraer_edicion(update.message.text.strip())
            monto = datos.get("monto", 0)
            descripcion = datos.get("descripcion", "")
            categoria = datos.get("categoria", "Otros")

            if not monto or float(monto) == 0:
                await update.message.reply_text(
                    "⚠️ No pude detectar el monto. Intenta de nuevo, ej:\n_\"50 comida de gato\"_",
                    parse_mode="Markdown"
                )
                edicion_pendiente[tid] = {"id": trans_id}
                return

            editar_transaccion(trans_id, obtener_o_crear_usuario(tid), monto, descripcion, categoria)

            # Preguntar medio de pago
            teclado_medio = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("📱 Yape", callback_data=f"edit_medio_{trans_id}_Yape"),
                    InlineKeyboardButton("💙 Plin", callback_data=f"edit_medio_{trans_id}_Plin"),
                ],
                [
                    InlineKeyboardButton("💵 Efectivo", callback_data=f"edit_medio_{trans_id}_Efectivo"),
                    InlineKeyboardButton("💳 Tarjeta", callback_data=f"edit_medio_{trans_id}_Tarjeta"),
                ],
                [
                    InlineKeyboardButton("🏦 Transferencia", callback_data=f"edit_medio_{trans_id}_Transferencia"),
                    InlineKeyboardButton("⏭️ Mantener", callback_data=f"edit_medio_{trans_id}_skip"),
                ],
            ])
            emoji = EMOJIS_CATEGORIA.get(categoria, "📦")
            await update.message.reply_text(
                f"✅ *Transacción #{trans_id} actualizada*\n\n"
                f"💵 S/ {float(monto):.2f}\n"
                f"📝 {descripcion}\n"
                f"{emoji} {categoria}\n\n"
                f"¿Actualizar el medio de pago?",
                parse_mode="Markdown",
                reply_markup=teclado_medio
            )
        except Exception as e:
            await update.message.reply_text(
                "⚠️ No pude interpretar los datos. Intenta de nuevo, ej:\n_\"50 comida de gato\"_",
                parse_mode="Markdown"
            )
            edicion_pendiente[tid] = {"id": trans_id}
        return
    # Pago fijo pendiente desde botón
    if tid in pago_fijo_pendiente:
        datos = pago_fijo_pendiente[tid]
        if "descripcion" not in datos:
            datos["descripcion"] = update.message.text.strip()
            await update.message.reply_text("💵 ¿Cuánto es el monto? _(ej: 99)_", parse_mode="Markdown")
        elif "monto" not in datos:
            try:
                datos["monto"] = float(update.message.text.strip().replace(",", "."))
                await update.message.reply_text("📅 ¿Qué día del mes vence? _(del 1 al 31)_")
            except ValueError:
                await update.message.reply_text("⚠️ Ingresa solo el número.")
        else:
            try:
                dia = int(update.message.text.strip())
                if not 1 <= dia <= 31:
                    raise ValueError
                usuario_id = obtener_o_crear_usuario(tid)
                categoria = clasificar_gasto(datos["descripcion"])
                guardar_pago_fijo(usuario_id, datos["descripcion"], datos["monto"], dia, categoria)
                pago_fijo_pendiente.pop(tid)
                await update.message.reply_text(
                    f"✅ Pago fijo registrado: {datos['descripcion']} — S/ {datos['monto']:.2f} el día {dia}",
                    reply_markup=menu_principal()
                )
            except ValueError:
                await update.message.reply_text("⚠️ Ingresa un número entre 1 y 31.")
        return
    # Registro manual pendiente
    if tid in registro_manual_pendiente:
        await handle_descripcion_manual(update, context)
        return
    await handle_texto(update, context)


if __name__ == "__main__":
    import sys
    import asyncio
    import signal
    if sys.platform == "win32":
        # Python 3.14+: WindowsSelectorEventLoopPolicy deprecado
        # Usar ProactorEventLoop con manejo manual del loop
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
    else:
        # En Linux/nube: manejar SIGTERM para apagado limpio
        signal.signal(signal.SIGTERM, lambda s, f: sys.exit(0))
    try:
        main()
    except KeyboardInterrupt:
        print("\nBot detenido.")
    except SystemExit:
        pass
